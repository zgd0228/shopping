import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

# 设置pandas对象全部显示
pd.set_option('display.max_columns', 1000)
pd.set_option('display.width', 1000)
pd.set_option('display.max_colwidth', 1000)


def get_df():
    """
    生成目标pandas对象
    生成一个，shape是，（1２，２）的pandas对象，列索引是多重索引，其他三列是计算列，以后再加
    :return:
    """
    # axis=0索引
    indexs = [
        'aaa', 'bbb', 'ccc', 'ddd', 'eee', 'fff', 'ggg',
        'hhh', 'iii', 'jjj', 'kkk', 'lll'
    ]
    # axis = 1 是一个多层索引
    column_index_level_1 = ['CC'] * 2
    column_index_level_2 = ['AA', 'BB']
    # 多层索引，可以用多维数组创建
    column_all = [column_index_level_1, column_index_level_2]

    # 数据准备
    df = pd.DataFrame(data=np.random.rand(12, 2), index=indexs, columns=column_all)

    # 将数据扩大100万倍
    df = df * 100000
    return df


def add_column_and_row(df):
    """
    添加行和列，并且格式化百分位的数据
    :param df:
    :return:
    """
    # 添加 CC, 'AA-BB' 'AA+BB' 'AA/BB' 列
    df[('CC', 'AA-BB')] = df.apply(lambda x: x[('CC', 'AA')] - x[('CC', 'BB')], axis=1)
    df[('CC', 'AA+BB')] = df.apply(lambda x: x[('CC', 'AA')] + x[('CC', 'BB')], axis=1)
    df[('CC', 'AA/BB')] = df.apply(lambda x: x[('CC', 'AA')] / x[('CC', 'BB')] if x[('CC', 'BB')] != 0 else np.NAN, axis=1)

    # 添加 total行
    df.loc['total'] = df.apply(lambda x: x.sum(), axis=0)

    # AA/BB 格式化，并且保留两位小数
    df[('CC', 'AA/BB')] = df[('CC', 'AA/BB')].apply(lambda x: format(x, '.2%'))

def generate_excel(df, excel_name='E:/test.xlsx', sheet_name='total'):
    """
    生成excel
    :param sheet_name:
    :param excel_name:
    :param df:
    :return:
    """
    # 如果文件存在，直接添加sheet；不存在则创建新的文件
    excel_writer = pd.ExcelWriter(excel_name, engine='openpyxl')
    if not os.path.exists(excel_name):
        df.to_excel(excel_writer, sheet_name=sheet_name)
    else:
        book = openpyxl.load_workbook(excel_writer.path)
        excel_writer.book = book
    df.to_excel(excel_writer, sheet_name=sheet_name)
    excel_writer.save()
    excel_writer.close()

def set_excel_style(df, excel_name='E:/test.xlsx', sheet_name='total'):
    book = openpyxl.load_workbook(excel_name)
    sheet = book[sheet_name]



    # 格式化数字
    for r, row in enumerate(sheet.rows):
        for c, cell in enumerate(row):
            if cell.row > 6:
                cell.number_format = '#,##0;-#,##0'
                cell.alignment = Alignment(horizontal='left', vertical='center')

    border_none = Border(top=Side(), bottom=Side(), left=Side(), right=Side())
    border_double = Border(top=Side(border_style='thin'), bottom=Side(border_style='double'), left=Side(), right=Side())
    # 设置整体的字体
    for row in sheet.rows:
        for cell in row:
            cell.border = border_none

    # 前四行的第一个cell，加粗靠左
    for cell in sheet['A']:
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if cell.value == 'total':
            for cell__ in sheet[cell.row]:
                if cell__.column > 1:
                    cell__.border = border_double

    # 合并列
    sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=6)



    # 列折叠
    sheet.column_dimensions.group('B', 'D', hidden=True)

    # 行折叠
    sheet.row_dimensions.group(7, 7 + len(df.index) - 2, hidden=True)
    book.save(excel_name)


if __name__ == '__main__':
    df = get_df()
    add_column_and_row(df)
    generate_excel(df)
    set_excel_style(df)
