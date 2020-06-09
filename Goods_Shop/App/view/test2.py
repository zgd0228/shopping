import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os


# 设置pandas对象全部显示
# pd.set_option('display.max_columns', 1000)
# pd.set_option('display.width', 1000)
# pd.set_option('display.max_colwidth', 1000)

# book = openpyxl.load_workbook('E:/test2.xlsx')
# sheet = book['Sheet1']
#
# sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
# font = Font(name=u'宋体', bold = True)
# align = Alignment(horizontal='center', vertical='center')
#
# sheet.cellstyle('A1', font, align)
# sheet.cell(1,1).value='合并'
# book.save('E:/test2.xlsx')


data_list = [[1,2,3,4,5,6,7,8,9,10,11,12]]
data_df = pd.DataFrame(data=data_list)
path = 'E:/ecd.xlsx'
header = ['累计','当日新增','当月新增','累计','当日新增','当月新增','累计','当日新增','当月新增','累计','当日新增','当月新增']
data_df.to_excel(path,index=False,header=header,startrow=3)

rows = len(data_df)
wb=load_workbook(path)

ws=wb.active
cols=['A','B','C','D','E','F','G','H','I','J','K','L']   # 需要处理的列
widths=[12,12,12,12,12,12,12,12,12,12,12,12]   # 各列的列宽

ws.cell(row=1,column=1,value='%s年重庆分行IC卡发卡明细表')  # 设置第一行标题
ws.cell(row=2,column=1,value='2020-1-1')   # 写入制卡量总数
ws.cell(row=3,column=1,value='开户统计')   # 写入制卡量总数
ws.cell(row=3,column=4,value='营销统计')   # 写入制卡量总数
ws.cell(row=3,column=7,value='有价值统计')   # 写入制卡量总数
ws.cell(row=3,column=10,value='存款统计')   # 写入制卡量总数

for i in range(12):
  ws.column_dimensions[cols[i]].width=widths[i]   # 设置列宽

for i in range(3,rows+4):
  ws.row_dimensions[i].height=15   # 设置行高

ws.merge_cells('A1:L1')   # 合并第一行的单元格
ws.merge_cells('A2:L2')   # 合并第一行的单元格
ws.merge_cells('A3:C3')   # 合并第一行的单元格
ws.merge_cells('D3:F3')   # 合并第一行的单元格
ws.merge_cells('G3:I3')   # 合并第一行的单元格
ws.merge_cells('J3:L3')   # 合并第一行的单元格

ws['A1'].font=Font(size=18,bold=True)   # 设置合并单元格的字体大小、加粗

ws['D3'].font=Font(size=12,bold=True)   # 设置合并单元格的字体大小、加粗
ws['G3'].font=Font(size=12,bold=True)   # 设置合并单元格的字体大小、加粗
ws['J3'].font=Font(size=12,bold=True)   # 设置合并单元格的字体大小、加粗
for i in cols:
  ws[i+'3'].font=Font(size=14,bold=True)  # 设置第二行字体大小、加粗
for i in range(3,rows+3):
  ws['A'+str(i)].font=Font(size=12)  # 设置日期列的字体大小

for i in cols:
  for j in range(1,rows+5):
    ws[i+str(j)].alignment=Alignment(horizontal='center',vertical='center')   # 设置上下左右居中


for i in cols:
  for j in range(2,rows+5):
    ws[i+str(j)].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))   # 设置单元格的上下左右为细边框
ws['A2'].font=Font(size=12,bold=False)   # 设置合并单元格的字体大小、加粗
ws['A3'].font=Font(size=12,bold=True)   # 设置合并单元格的字体大小、加粗
wb.save(path)   # 保存设置










